import os
import sys
import datetime

import jetson.inference
import jetson.utils

import threading

import cv2 #camera
import pyimgur #upload photo
#import pygsheets #fetch data from google sheet
from openpyxl import load_workbook #excel
from openpyxl import Workbook

from argparse import ArgumentParser
from datetime import datetime
from glob import glob
from time import sleep

from flask import Flask, request, abort, Response, send_from_directory
from linebot import (
    LineBotApi, WebhookHandler 
)
from linebot.exceptions import (
    InvalidSignatureError
)
from linebot.models import *


num=0

net = jetson.inference.detectNet("pednet", threshold=0.5) #choose ped-100 model
camera = jetson.utils.videoSource("/dev/video0")      # choose webcam
display = jetson.utils.videoOutput("room.jpg") # choose image as output
num=0

#object detect
def detect():
    time = 70
    while display.IsStreaming():
        img = camera.Capture()
        detections = net.Detect(img)
        global num
        num=len(detections)

        #constraint that it will only send one warning message every 15 minutes 
        if time != 70 and abs(datetime.now().minute-time) > 15:
            time=70
        if num > 2  and time == 70:
            list=[]
            r=1
            wb = load_workbook('id.xlsx') #create a file
            sheet = wb.active
            while sheet.cell(row = r, column = 1).value != None:
                if sheet.cell(row = r, column = 3).value == None:
                    list.append(sheet.cell(row = r, column = 1).value)
                r=r+1
            if len(list) > 0:
                time = datetime.now().minute
                line_bot_api.multicast(list, TextSendMessage(text='現場人數已超過2人,請注意您的安全!\n如果有必要，請自行離開!\n管理員關心您!'))

        #output setting
        display.Render(img)
        display.SetStatus("Object Detection | Network {:.0f} FPS".format(net.GetNetworkFPS()))

        #sleep to lightweight the program 
        sleep(1)

#setting muti-thread
t=threading.Thread(target=detect)
t.start()



#setting Line Bot information
app = Flask(__name__)
channel_secret = "2c5958918a34c00a0ccd046d2ac34bc6"
channel_access_token = "wWc7YMXxvxILzxreW6fzHA1BOh8usaeeWI4iy2GQpbpvGn4XIK8s1jF2F3cSaCpOCLLS9ZP5LYZWTkXPHXfSlMMmc7ChziFzkKfjPiPsBh9LULr3PE5q1+VSdk9ngp+79RhHB9lJIBd4LGc5uU/L/QdB04t89/1O/w1cDnyilFU="

if channel_secret == "":
    print("Please specify LINE_CHANNEL_SECRET.")
    sys.exit(1)
if channel_access_token == "":
    print("Please specify LINE_CHANNEL_ACCESS_TOKEN.")
    sys.exit(1)

line_bot_api = LineBotApi(channel_access_token)
handler = WebhookHandler(channel_secret)

#push message to one manager
line_bot_api.push_message('U0c91f337f97510e29e594415dc135cda', 
    TextSendMessage(text='歡迎光臨！'))

#webhook
@app.route("/callback", methods=["POST"])
def callback():
    # get X-Line-Signature header value
    signature = request.headers["X-Line-Signature"]

    # get request body as text
    body = request.get_data(as_text=True)
    app.logger.info("Request body: " + body)
    
    # handle webhook body
    try:
        handler.handle(body, signature)
    except InvalidSignatureError:
        abort(400)

    return "OK"


#Message API
@handler.add(MessageEvent, message=TextMessage)
def handle_text_message(event):
    text = event.message.text
    global num
    if text == '現在有多少人？':
                
        #bot reply
        line_bot_api.reply_message(
                event.reply_token, 
                    TextSendMessage(text='現場有' + str(num) + '個人'),
                    #TextSendMessage(text='Status message: ' + str(profile.status_message))
                
        )
    elif text == "歡迎!":
        user_id = event.source.user_id #get user id
        t = datetime.now() #get time that the user enters
        wb = load_workbook('id.xlsx') #create a file
        sheet = wb.active
        sheet.append([user_id, t])
        wb.save("id.xlsx")

    elif text == "再見!":
        r = 1
        user_id = event.source.user_id #get user id
        t = datetime.now() #get time that the user enters
        wb = load_workbook('id.xlsx') #create a file
        sheet = wb.active
        while sheet.cell(row = r, column = 1).value != user_id or sheet.cell(row = r, column = 3).value != None:
            r = r + 1
        sheet.cell(row = r, column = 3, value = t)
        wb.save("id.xlsx")
       
        
    elif text == "請看鏡頭":
        


        #show photo url
        CLIENT_ID = "1a91db4a650ce5c"
        PATH = "room.jpg" #a filepath to an image on your computer
        title = "Uploaded with PyImgur"
        im = pyimgur.Imgur(CLIENT_ID)
        uploaded_image = im.upload_image(PATH, title=title)
        print(uploaded_image.link)
        
        #bot reply
       	line_bot_api.reply_message(
            event.reply_token, ImageSendMessage(
		original_content_url=uploaded_image.link,
				preview_image_url=uploaded_image.link		
            )
        )
    elif text.lower() == "menu":
        #bot reply
        line_bot_api.reply_message(
                event.reply_token, 
                    TemplateSendMessage(
                    alt_text = 'Welcome',
                    template = ButtonsTemplate(
                    thumbnail_image_url = "https://dep.hcchb.gov.tw/userfiles/4307/images/%E5%B8%82%E5%BA%9C%E9%98%B2%E7%96%AB%E5%A4%A7%E4%BD%9C%E6%88%B0-1.jpg",
                    title = "防疫專區",
                    text = "齊心抗疫，台灣加油",
                    actions = [

                        MessageTemplateAction(
                        label = "進入",
                        text = "歡迎!"
                        ),
                        MessageTemplateAction(
                        label = "查看人數",
                        text = "現在有多少人？"
                        ),
                        MessageTemplateAction(
                        label = "拍照",
                        text = "請看鏡頭"
                        ),
                        MessageTemplateAction(
                        label = "離開",
                        text = "再見!"
                        )
                    ]
                )
            )
        )

    #echo
    else:
        line_bot_api.reply_message( 
                event.reply_token, 
                    TextSendMessage(text = event.message.text)
        )
       
if __name__ == "__main__":
    arg_parser = ArgumentParser(
        usage="Usage: python3 " + __file__ + " [--port <port>] [--help]"
    )
    arg_parser.add_argument("-p", "--port", default=8000, help="port")
    arg_parser.add_argument("-d", "--debug", default=False, help="debug")
    options = arg_parser.parse_args()

    app.run(debug=options.debug, port=options.port)
